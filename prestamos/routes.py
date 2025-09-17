from __future__ import annotations
import os
from datetime import date, datetime
from flask import (
    request,
    render_template,
    jsonify,
    send_file,
    url_for,
    current_app,
    flash,
    redirect,
)
from .services import (
    generar_cronograma,
    nombre_mes,
    PDF_CSS,
    amortizar,
    dec,
    nombre_empleado,
)
from weasyprint import HTML, CSS
from decimal import Decimal, ROUND_HALF_UP

from sqlalchemy import and_, or_, func
from flask_login import login_required

from . import prestamos_bp
from .models import Prestamo, Cuota, Documento
from .services import generar_cronograma, nombre_mes, PDF_CSS, amortizar, dec
from models import db, Empleado

#!#######################################ARREGLO DE VISUALIZACION DATA EN FORMHTML##################################################


@prestamos_bp.get("/prestamos")
@login_required
def prestamos_index():
    dni = (request.args.get("dni") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    prestamos = []
    pagination = None
    page_totals = {"monto": 0.0, "saldo": 0.0}
    totals_dni = None

    if len(dni) == 8 and dni.isdigit():
        q = Prestamo.query.filter(
            Prestamo.dni == dni, Prestamo.estado != "Cancelado"
        ).order_by(Prestamo.id.asc())
        pagination = q.paginate(page=page, per_page=per_page, error_out=False)
        prestamos = pagination.items

        total_monto, total_saldo = (
            db.session.query(
                func.coalesce(func.sum(Prestamo.monto), 0.0),
                func.coalesce(func.sum(Prestamo.saldo), 0.0),
            )
            .filter(Prestamo.dni == dni)
            .one()
        )
        totals_dni = {
            "monto": float(total_monto or 0),
            "saldo": float(total_saldo or 0),
        }
    else:
        sub = (
            db.session.query(Prestamo.id)
            .filter(Prestamo.estado != "Cancelado")  # ⬅️ aquí
            .order_by(Prestamo.id.desc())
            .limit(20)
            .subquery()
        )
        prestamos = (
            db.session.query(Prestamo)
            .join(sub, Prestamo.id == sub.c.id)
            .order_by(Prestamo.id.asc())
            .all()
        )

    page_totals["monto"] = sum(float(p.monto or 0) for p in prestamos)
    page_totals["saldo"] = sum(float(p.saldo or 0) for p in prestamos)

    return render_template(
        "prestamos/index.html",
        prestamos=prestamos,
        pagination=pagination,
        dni=dni,
        per_page=per_page,
        page_totals=page_totals,
        totals_dni=totals_dni,
    )


@prestamos_bp.route("/prestamos/nuevo")
def ui_nuevo_prestamo():
    dni = request.args.get("dni", "").strip()
    empleado = Empleado.query.filter_by(dni=dni).first() if dni else None
    return render_template("prestamos/form.html", empleado=empleado, hoy=date.today())


@prestamos_bp.route("/api/prestamos/cronograma/preview", methods=["POST"])
def api_preview_cronograma():
    d = request.get_json(force=True)
    try:
        items = generar_cronograma(
            d.get("monto_total"),
            int(d.get("n_cuotas")),
            int(d.get("mes_inicio")),
            int(d.get("anio_inicio")),
            bool(d.get("incluir_grati")),
            int(d.get("anio_grati_desde")) if d.get("anio_grati_desde") else None,
        )
        # 👇 convertir Decimal -> float para JSON
        for it in items:
            it["monto"] = float(it["monto"])
        total = round(sum(it["monto"] for it in items), 2)
        return jsonify({"items": items, "total": total})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


def _normalizar_cuotas_custom(raw_items, monto_total, n_cuotas):
    """Valida/normaliza cuotas_custom y ajusta la última para cuadrar el total."""
    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError("cuotas_custom inválido")
    items = []
    for i, it in enumerate(raw_items, start=1):
        orden = int(it.get("orden") or i)
        etiqueta = (it.get("etiqueta") or f"Cuota {i}").strip()
        anio = int(it.get("anio") or 0)
        mes = int(it.get("mes") or 0)
        es_grati = bool(it.get("es_grati"))
        monto = dec(it.get("monto") or 0).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if mes and not (1 <= mes <= 12):
            raise ValueError(f"Mes inválido en cuota #{i}")
        items.append(
            {
                "orden": orden,
                "etiqueta": etiqueta,
                "anio": anio,
                "mes": mes,
                "es_grati": es_grati,
                "monto": monto,
            }
        )
    if len(items) != int(n_cuotas):
        raise ValueError("El número de cuotas no coincide con n_cuotas")

    # Ajuste final para cuadrar total
    M = dec(monto_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    suma = sum(dec(x["monto"]) for x in items).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    diff = (M - suma).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    items[-1]["monto"] = (dec(items[-1]["monto"]) + diff).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return items


@prestamos_bp.route("/api/prestamos", methods=["POST"])
def api_crear_prestamo():
    d = request.get_json(force=True)
    emp = Empleado.query.filter_by(dni=(d.get("dni") or "").strip()).first()
    if not emp:
        return jsonify({"error": "Colaborador no existe. Registre primero."}), 400

    incluir_grati = bool(d.get("incluir_grati"))
    anio_grati_desde = (
        int(d.get("anio_grati_desde")) if d.get("anio_grati_desde") else None
    )
    if incluir_grati and not anio_grati_desde:
        return jsonify({"error": "Debe indicar 'Año desde' para gratificaciones"}), 400

    p = Prestamo(
        empleado_id=emp.id,
        tipo=d.get("tipo"),
        motivo_especifico=d.get("motivo_especifico"),
        fecha_solicitud=datetime.strptime(d.get("fecha_solicitud"), "%Y-%m-%d").date(),
        monto_total=d.get("monto_total"),
        n_cuotas=int(d.get("n_cuotas")),
        incluir_grati=incluir_grati,
        anio_grati_desde=anio_grati_desde,
        fecha_firma=datetime.strptime(d.get("fecha_firma"), "%Y-%m-%d").date(),
        estado="Emitido",
        creado_por=(d.get("usuario") or "web"),
    )
    db.session.add(p)
    db.session.flush()

    # Cuotas
    raw_custom = d.get("cuotas_custom")
    if raw_custom:
        try:
            items = _normalizar_cuotas_custom(
                raw_custom, d.get("monto_total"), int(d.get("n_cuotas"))
            )
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
    else:
        items = generar_cronograma(
            d.get("monto_total"),
            int(d.get("n_cuotas")),
            int(d.get("mes_inicio")),
            int(d.get("anio_inicio")),
            incluir_grati,
            anio_grati_desde,
        )

    for it in items:
        try:
            fct = (
                date(int(it["anio"]), int(it["mes"]), 1)
                if it.get("anio") and it.get("mes")
                else None
            )
        except Exception:
            fct = None
        db.session.add(
            Cuota(
                prestamo_id=p.id,
                orden=int(it["orden"]),
                etiqueta=it["etiqueta"],
                anio=int(it.get("anio") or 0),
                mes=int(it.get("mes") or 0),
                es_grati=bool(it.get("es_grati")),
                monto=dec(it["monto"]),
                fecha_cobro_teorica=fct,
            )
        )

    db.session.commit()

    # 🔹 Construcción del nombre de archivo
    fecha_str = p.fecha_firma.strftime("%Y-%m-%d")
    tipo_str = (p.tipo or "OTROS").upper().replace(" ", "_")
    nombre_str = (nombre_empleado(emp) or "").upper().replace(" ", "_")
    filename = f"{fecha_str}_DESCUENTO_{tipo_str}_{nombre_str}.pdf"

    return jsonify(
        {
            "id": p.id,
            "pdf_url": url_for(
                "prestamos.pdf_prestamo", prestamo_id=p.id, _external=True
            ),
            "filename": filename,
        }
    )


@prestamos_bp.route("/prestamos/<int:prestamo_id>/pdf")
def pdf_prestamo(prestamo_id: int):
    import os, unicodedata

    p = Prestamo.query.get_or_404(prestamo_id)

    # ---- Helper para limpiar y poner en MAYÚSCULAS con '_' ----
    def _slug_upper(s: str) -> str:
        if not s:
            return ""
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))  # quita tildes
        s = s.upper().replace(" ", "_")
        s = "".join(
            ch for ch in s if ch.isalnum() or ch in ("_", "-")
        )  # seguro p/archivo
        return s

    emp = p.empleado
    emp_nombre = nombre_empleado(emp)

    html = render_template(
        "prestamos/pdf.html",
        p=p,
        emp=emp,
        emp_nombre=emp_nombre,
        cuotas=p.cuotas,
        nombre_mes=nombre_mes,
        hoy=date.today(),
    )
    pdf = HTML(string=html, base_url=request.url_root).write_pdf(
        stylesheets=[CSS(string=PDF_CSS)]
    )

    # ---- Construcción del nombre final ----
    fecha_str = (
        p.fecha_firma.strftime("%Y-%m-%d")
        if p.fecha_firma
        else date.today().strftime("%Y-%m-%d")
    )
    tipo_str = _slug_upper(p.tipo or "OTROS")
    nombre_str = _slug_upper(emp_nombre or "")
    filename = f"{fecha_str}_DESCUENTO_{tipo_str}_{nombre_str}.pdf"

    # ---- Guardar en disco con ese nombre ----
    ruta = os.path.join("storage", "prestamos")
    os.makedirs(ruta, exist_ok=True)
    fullpath = os.path.join(ruta, filename)
    with open(fullpath, "wb") as f:
        f.write(pdf)

    # ---- Persistir registro del documento ----
    db.session.add(Documento(prestamo_id=p.id, ruta_pdf=fullpath))
    db.session.commit()

    # ---- Enviar con nombre de descarga correcto ----
    return send_file(fullpath, as_attachment=True, download_name=filename)


@prestamos_bp.route("/prestamos/export-excel")
def export_excel():
    import os
    import pandas as pd
    from pandas import ExcelWriter
    from datetime import date as _date
    from flask import send_file
    from openpyxl.utils import column_index_from_string, get_column_letter

    # ------------------ Datos base ------------------
    Qp = (
        db.session.query(Prestamo, Empleado)
        .join(Empleado, Prestamo.empleado_id == Empleado.id)
        .all()
    )
    rows_p = [
        {
            "ID_PRESTAMO": p.id,
            "DNI": e.dni,
            "NOMBRE": nombre_empleado(e),
            "PRESTAMO": (
                p.tipo if p.tipo != "Otros" else f"Otros: {p.motivo_especifico or ''}"
            ),
            "MONTO TOTAL": float(p.monto_total),
            "FECHA DE SOLICITUD": p.fecha_solicitud.strftime("%Y-%m-%d"),
            "AÑO": p.fecha_solicitud.year,
            "MONTO DE AMORTIZACIÓN": (
                float(sum(a.monto for a in p.amortizaciones))
                if p.amortizaciones
                else 0.0
            ),
            "FECHA DE AMORTIZACIÓN": max(
                (a.fecha for a in p.amortizaciones), default=None
            ),
            "OBS DE AMORTIZACIÓN": (
                "; ".join(filter(None, [a.observacion for a in p.amortizaciones]))
                if p.amortizaciones
                else ""
            ),
        }
        for p, e in Qp
    ]

    Qc = (
        db.session.query(Cuota, Prestamo, Empleado)
        .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
        .join(Empleado, Prestamo.empleado_id == Empleado.id)
        .all()
    )
    rows_c = [
        {
            "ID_PRESTAMO": p.id,
            "DNI": e.dni,
            "FECHA_COBRO": (
                _date(c.anio, c.mes, 1).strftime("%Y-%m-%d") if c.anio and c.mes else ""
            ),
            "ETIQUETA": c.etiqueta,
            "MONTO": float(c.monto),
            "ESTADO": (c.estado or "Pendiente").strip(),
        }
        for c, p, e in Qc
    ]

    df_p = pd.DataFrame(rows_p)
    df_c = pd.DataFrame(rows_c)

    # ------------------ Cronograma en columnas (MISMA hoja) ------------------
    month_cols = []
    if not df_c.empty:
        df_c["ESTADO"] = df_c["ESTADO"].fillna("Pendiente").astype(str)
        mask_amort = df_c["ESTADO"].str.strip().str.lower().eq("amortizada")
        df_c.loc[mask_amort, "MONTO"] = 0.0

        MES_ABBR = {
            1: "ene",
            2: "feb",
            3: "mar",
            4: "abr",
            5: "may",
            6: "jun",
            7: "jul",
            8: "ago",
            9: "set",
            10: "oct",
            11: "nov",
            12: "dic",
        }

        df_c["FECHA_COBRO"] = pd.to_datetime(df_c["FECHA_COBRO"], errors="coerce")
        df_c["is_grati"] = df_c["ETIQUETA"].str.contains("grat", case=False, na=False)

        df_c["col"] = df_c["FECHA_COBRO"].apply(
            lambda d: (
                f"{MES_ABBR.get(d.month,'')} {str(d.year)[-2:]}"
                if pd.notnull(d)
                else None
            )
        )
        df_c.loc[df_c["is_grati"], "col"] = df_c.loc[
            df_c["is_grati"], "FECHA_COBRO"
        ].apply(
            lambda d: (
                f"grati {MES_ABBR.get(d.month,'')} {str(d.year)[-2:]}"
                if pd.notnull(d)
                else None
            )
        )

        today = pd.Timestamp.today().normalize()

        def sort_key(r):
            d = r["FECHA_COBRO"]
            if pd.isna(d):
                return 9e9
            months = (d.year - today.year) * 12 + (d.month - today.month)
            return months + (-0.5 if r["is_grati"] else 0.0)

        df_c["sort_key"] = df_c.apply(sort_key, axis=1)
        col_order = (
            df_c.loc[df_c["col"].notna(), ["col", "sort_key"]]
            .drop_duplicates()
            .sort_values("sort_key")
        )["col"].tolist()

        pivot_mes = (
            pd.pivot_table(
                df_c,
                index=["ID_PRESTAMO"],
                columns=["col"],
                values="MONTO",
                aggfunc="sum",
            )
            .reindex(columns=col_order)
            .fillna(0.0)
            .reset_index()
        )

        df_p = df_p.merge(pivot_mes, on="ID_PRESTAMO", how="left")
        month_cols = [c for c in col_order if c in df_p.columns]
        if month_cols:
            df_p[month_cols] = df_p[month_cols].fillna(0.0)
            base_cols = list(df_p.columns)
            base_cols_wo = [c for c in base_cols if c not in month_cols]
            pos = base_cols_wo.index("AÑO") + 1
            new_order = base_cols_wo[:pos] + month_cols + base_cols_wo[pos:]
            df_p = df_p[new_order]

    # ------------------ Pivot extra (opcional) ------------------
    pivot = (
        pd.pivot_table(
            df_c, index=["DNI"], columns=["ETIQUETA"], values="MONTO", aggfunc="sum"
        )
        .fillna(0)
        .reset_index()
        if not df_c.empty
        else pd.DataFrame()
    )

    # ------------------ Escribir Excel + FORMATO numérico ------------------
    os.makedirs("storage/exports", exist_ok=True)
    out_path = os.path.join("storage/exports", "Prestamos.xlsx")
    with ExcelWriter(out_path, engine="openpyxl") as w:
        df_p.to_excel(w, index=False, sheet_name="Prestamos")
        if not pivot.empty:
            pivot.to_excel(w, index=False, sheet_name="Reporte_Pivot")

        # ===== Formato en hoja Prestamos =====
        ws = w.sheets["Prestamos"]

        # Mapa encabezado -> letra de columna
        header_to_col_letter = {cell.value: cell.column_letter for cell in ws[1]}

        # Columnas a formatear (numéricas)
        targets = ["MONTO TOTAL", "MONTO DE AMORTIZACIÓN"] + month_cols

        for header in targets:
            col_letter = header_to_col_letter.get(header)
            if not col_letter:
                continue

            col_idx = column_index_from_string(col_letter)
            # Aplica formato a TODA la columna desde fila 2 a max_row
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                cell = row[0]
                # Si vino como texto por alguna razón, intenta convertir
                if isinstance(cell.value, str):
                    try:
                        cell.value = float(cell.value)
                    except Exception:
                        pass
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

            # Ajuste de ancho
            ws.column_dimensions[col_letter].width = max(12, len(header) + 2)

        # Autofiltro y panes inmovilizados (ayuda de lectura)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # ===== Formato en hoja Reporte_Pivot (si existe) =====
        if "Reporte_Pivot" in w.sheets:
            ws2 = w.sheets["Reporte_Pivot"]
            ws2.auto_filter.ref = ws2.dimensions
            ws2.freeze_panes = "A2"
            # Formatear todas las columnas numéricas excepto la 1 (DNI)
            for j, cell in enumerate(ws2[1], start=1):
                if j == 1:
                    continue
                col_letter = cell.column_letter
                col_idx = column_index_from_string(col_letter)
                for row in ws2.iter_rows(
                    min_row=2, max_row=ws2.max_row, min_col=col_idx, max_col=col_idx
                ):
                    c = row[0]
                    if isinstance(c.value, str):
                        try:
                            c.value = float(c.value)
                        except Exception:
                            pass
                    if isinstance(c.value, (int, float)):
                        c.number_format = "#,##0.00"
                ws2.column_dimensions[col_letter].width = max(
                    12, len(str(cell.value)) + 2
                )

    return send_file(out_path, as_attachment=True, download_name="Prestamos.xlsx")


@prestamos_bp.route("/api/prestamos")
def api_listar_prestamos():
    dni = (request.args.get("dni") or "").strip()
    limit = request.args.get("limit", type=int)

    base = db.session.query(Prestamo, Empleado).join(
        Empleado, Prestamo.empleado_id == Empleado.id
    )

    if dni:
        q = base.filter(Empleado.dni == dni, Prestamo.estado != "Cancelado").order_by(Prestamo.id.asc())
    else:
        n = limit or 20
        sub = (
            db.session.query(Prestamo.id)
            .filter(Prestamo.estado != "Cancelado")
            .order_by(Prestamo.id.desc())
            .limit(n)
            .subquery()
        )
        q = base.join(sub, Prestamo.id == sub.c.id).order_by(Prestamo.id.asc())

    rows = q.all()

    data = []
    for p, e in rows:
        saldo = float(
            sum(c.monto for c in p.cuotas if (c.estado or "Pendiente") == "Pendiente")
        )
        data.append(
            {
                "id": p.id,
                "dni": e.dni,
                "nombre": nombre_empleado(e),
                "tipo": p.tipo,
                "monto_total": float(p.monto_total),
                "saldo_pendiente": round(saldo, 2),
                "estado": p.estado,
                "fecha_solicitud": p.fecha_solicitud.strftime("%Y-%m-%d"),
            }
        )
    return jsonify(data)


@prestamos_bp.route("/api/prestamos/<int:prestamo_id>/amortizacion", methods=["POST"])
def api_amortizacion(prestamo_id: int):
    d = request.get_json(force=True)
    p = Prestamo.query.get_or_404(prestamo_id)
    monto = d.get("monto")
    fecha = datetime.strptime(d.get("fecha"), "%Y-%m-%d").date()
    amortizar(p, dec(monto), fecha, d.get("observacion"), d.get("usuario") or "web")
    db.session.commit()
    return jsonify({"ok": True, "estado": p.estado})


@prestamos_bp.route("/api/prestamos/<int:prestamo_id>/cuotas", methods=["GET"])
def api_cuotas_prestamo(prestamo_id: int):
    try:
        p = Prestamo.query.get_or_404(prestamo_id)
        cuotas = []
        for c in p.cuotas:
            anio = int(c.anio) if c.anio is not None else 0
            mes = int(c.mes) if c.mes is not None else 0
            fecha_cobro = f"{anio:04d}-{mes:02d}-01" if anio and mes else ""
            cuotas.append(
                {
                    "orden": int(c.orden or 0),
                    "etiqueta": c.etiqueta or "",
                    "monto": float(c.monto or 0),
                    "estado": (c.estado or "Pendiente").strip() or "Pendiente",
                    "es_grati": bool(c.es_grati),
                    "fecha_cobro": fecha_cobro,
                    "fecha_descuento_real": (
                        c.fecha_descuento_real.strftime("%Y-%m-%d")
                        if c.fecha_descuento_real
                        else None
                    ),
                }
            )
        saldo = round(sum(r["monto"] for r in cuotas if r["estado"] == "Pendiente"), 2)
        return jsonify(
            {
                "id": p.id,
                "dni": p.empleado.dni,
                "nombre": nombre_empleado(p.empleado),
                "tipo": p.tipo,
                "monto_total": float(p.monto_total or 0),
                "saldo_pendiente": saldo,
                "cuotas": cuotas,
            }
        )
    except Exception:
        current_app.logger.exception("Error en /api/prestamos/<id>/cuotas")
        return jsonify({"error": "Error interno"}), 500


####!SALDO#####
@prestamos_bp.route("/api/prestamos/<int:prestamo_id>/saldo", methods=["GET"])
def api_saldo_prestamo(prestamo_id: int):
    try:
        p = Prestamo.query.get_or_404(prestamo_id)
        # Misma regla que usas en /api/prestamos y /api/prestamos/<id>/cuotas
        saldo = round(
            sum(
                float(c.monto or 0)
                for c in p.cuotas
                if (c.estado or "Pendiente").strip() == "Pendiente"
            ),
            2,
        )
        return jsonify({"id": p.id, "saldo": saldo})
    except Exception:
        current_app.logger.exception("Error en /api/prestamos/<id>/saldo")
        return jsonify({"error": "Error interno"}), 500


@prestamos_bp.route("/api/prestamos/empleados", methods=["GET", "POST"])
def api_empleados():
    try:
        if request.method == "GET":
            dni = (request.args.get("dni") or "").strip()
            if not dni:
                return jsonify({"error": "DNI requerido"}), 400
            emp = Empleado.query.filter_by(dni=dni).first()
            if not emp:
                return jsonify([])
            return jsonify(
                [
                    {
                        "id": emp.id,
                        "dni": emp.dni,
                        "nombre": nombre_empleado(emp),
                        "cargo": getattr(emp, "cargo", None),
                        "direccion": getattr(emp, "direccion", None),
                    }
                ]
            )

        # POST (crear)
        data = request.get_json(force=True) or {}
        dni = (data.get("dni") or "").strip()
        if not dni:
            return jsonify({"error": "DNI requerido"}), 400

        existente = Empleado.query.filter_by(dni=dni).first()
        if existente:
            return jsonify({"id": existente.id, "mensaje": "Empleado ya existe"})

        emp = Empleado()

        def set_if(attr, value):
            """Setea solo si el atributo existe en el modelo."""
            if value in (None, ""):
                return
            if hasattr(Empleado, attr):
                setattr(emp, attr, value)

        # Campos de tu formulario de Convenios
        nombre_txt = (data.get("nombre") or data.get("nombre_completo") or "").strip()
        set_if("dni", dni)
        set_if("nombre", nombre_txt)
        set_if("nombre_completo", nombre_txt)
        set_if("cargo", (data.get("cargo") or "").strip())
        set_if("direccion", (data.get("direccion") or "").strip())

        # Compatibilidad con esquema legacy (si tienes nombres/apellidos)
        if nombre_txt and (
            hasattr(Empleado, "nombres") or hasattr(Empleado, "apellidos")
        ):
            partes = nombre_txt.split()
            if hasattr(Empleado, "nombres"):
                set_if(
                    "nombres", " ".join(partes[:-1]) if len(partes) > 1 else nombre_txt
                )
            if hasattr(Empleado, "apellidos"):
                set_if("apellidos", partes[-1] if len(partes) > 1 else "")

        # Fecha de ingreso (si existe en el modelo)
        fi = (data.get("fecha_ingreso") or "").strip()
        if fi and hasattr(Empleado, "fecha_ingreso"):
            from datetime import datetime as _dt

            try:
                emp.fecha_ingreso = _dt.strptime(fi, "%Y-%m-%d").date()
            except ValueError:
                return (
                    jsonify({"error": "fecha_ingreso inválida (use YYYY-MM-DD)"}),
                    400,
                )

        db.session.add(emp)
        db.session.commit()
        return jsonify({"id": emp.id})

    except Exception as e:
        current_app.logger.exception("Error en /api/prestamos/empleados")
        # Respuesta genérica para el front
        return jsonify({"error": "Error interno en empleados"}), 500


# ======= Helpers para borrar PDFs físicos =======
def _collect_pdf_paths(prestamo_id: int):
    try:
        docs = Documento.query.filter_by(prestamo_id=prestamo_id).all()
        return [d.ruta_pdf for d in docs if getattr(d, "ruta_pdf", None)]
    except Exception:
        current_app.logger.exception(
            "No se pudieron listar PDFs para préstamo %s", prestamo_id
        )
        return []


def _remove_files(paths):
    for pth in paths or []:
        try:
            if pth and os.path.exists(pth):
                os.remove(pth)
        except Exception:
            current_app.logger.warning("No se pudo borrar archivo %s", pth)


# ======= API: borrado duro =======
@prestamos_bp.route("/api/prestamos/<int:prestamo_id>", methods=["DELETE", "POST"])
def api_delete_prestamo(prestamo_id: int):
    p = Prestamo.query.get_or_404(prestamo_id)
    try:
        # 1) Captura rutas físicas ANTES de borrar Documentos
        file_paths = _collect_pdf_paths(p.id)

        # 2) Borra dependencias explícitas
        Cuota.query.filter_by(prestamo_id=p.id).delete(synchronize_session=False)
        Documento.query.filter_by(prestamo_id=p.id).delete(synchronize_session=False)
        try:
            from prestamos.models import Amortizacion  # opcional

            Amortizacion.query.filter_by(prestamo_id=p.id).delete(
                synchronize_session=False
            )
        except Exception:
            pass

        # 3) Borra el préstamo
        db.session.delete(p)
        db.session.commit()

        # 4) Borra PDFs del disco (fuera de la transacción)
        _remove_files(file_paths)

        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Error al eliminar préstamo %s", prestamo_id)
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500


# ======= (Opcional) UI por formulario =======
@prestamos_bp.route("/prestamos/<int:prestamo_id>/delete", methods=["POST"])
def ui_delete_prestamo(prestamo_id: int):
    p = Prestamo.query.get_or_404(prestamo_id)
    try:
        file_paths = _collect_pdf_paths(p.id)
        Cuota.query.filter_by(prestamo_id=p.id).delete(synchronize_session=False)
        Documento.query.filter_by(prestamo_id=p.id).delete(synchronize_session=False)
        try:
            from prestamos.models import Amortizacion

            Amortizacion.query.filter_by(prestamo_id=p.id).delete(
                synchronize_session=False
            )
        except Exception:
            pass
        db.session.delete(p)
        db.session.commit()
        _remove_files(file_paths)
        flash("Préstamo eliminado definitivamente.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Error al eliminar préstamo %s", prestamo_id)
        flash(f"No se pudo eliminar el préstamo. {type(e).__name__}: {e}", "danger")
    return redirect(request.referrer or url_for("prestamos.ui_nuevo_prestamo"))


# ==========================
# CIERRE DE PERIODOS ABIERTOS
# ==========================


@prestamos_bp.route("/api/prestamos/cerrar_mes", methods=["POST"])
def api_cerrar_mes():
    d = request.get_json(force=True) or {}
    try:
        anio = int(d.get("anio"))
        mes = int(d.get("mes"))
        if not (1 <= mes <= 12):
            raise ValueError("mes fuera de rango")
        fdesc = d.get("fecha_descuento") or date.today().isoformat()
        fecha_desc = datetime.strptime(fdesc, "%Y-%m-%d").date()
        dni = (d.get("dni") or "").strip()

        q = (
            db.session.query(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Empleado, Prestamo.empleado_id == Empleado.id)
            .filter(Cuota.estado == "Pendiente", Cuota.anio == anio, Cuota.mes == mes)
        )
        if dni:
            q = q.filter(Empleado.dni == dni)

        cuotas = q.all()
        total = Decimal("0.00")
        pids_tocados = set()

        for c in cuotas:
            c.estado = "Descontada"
            c.fecha_descuento_real = fecha_desc
            pids_tocados.add(c.prestamo_id)
            total += Decimal(c.monto)

        cancelados = 0
        parciales = 0
        for pid in pids_tocados:
            p = Prestamo.query.get(pid)
            if all(cc.estado != "Pendiente" for cc in p.cuotas):
                p.estado = "Cancelado"
                cancelados += 1
            else:
                p.estado = "Amortizado Parcial"
                parciales += 1

        db.session.commit()
        return jsonify(
            {
                "ok": True,
                "cerradas": len(cuotas),
                "monto": float(total.quantize(Decimal("0.01"))),
                "prestamos_cancelados": cancelados,
                "prestamos_parciales": parciales,
            }
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Fallo en /api/prestamos/cerrar_mes")
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500


# ==========================
# APERTURA DE PERIODOS CERRADOS
# ==========================


@prestamos_bp.route("/api/prestamos/aperturar_mes", methods=["POST"])
def api_aperturar_mes():
    """
    Reabre un mes: pasa a 'Pendiente' las cuotas con estado 'Descontada'
    para el mes/año indicado. Opcionalmente filtra por DNI y limpia la
    fecha de descuento real.

    Body JSON:
    - anio: int (obligatorio)
    - mes:  int (1..12, obligatorio)
    - dni:  str | null (opcional)
    - limpiar_fecha: bool (opcional, default True)
    """
    d = request.get_json(force=True) or {}
    try:
        anio = int(d.get("anio"))
        mes = int(d.get("mes"))
        if not (1 <= mes <= 12):
            raise ValueError("mes fuera de rango")

        dni = (d.get("dni") or "").strip()
        limpiar_fecha = (
            bool(d.get("limpiar_fecha")) if d.get("limpiar_fecha") is not None else True
        )

        # Solo reabrimos cuotas que fueron cerradas por planilla:
        q = (
            db.session.query(Cuota)
            .join(Prestamo, Cuota.prestamo_id == Prestamo.id)
            .join(Empleado, Prestamo.empleado_id == Empleado.id)
            .filter(
                Cuota.estado == "Descontada",
                Cuota.anio == anio,
                Cuota.mes == mes,
            )
        )
        if dni:
            q = q.filter(Empleado.dni == dni)

        cuotas = q.all()
        if not cuotas:
            return jsonify({"ok": True, "reabiertas": 0, "prestamos_afectados": 0})

        reabiertas = 0
        pids_tocados = set()

        for c in cuotas:
            c.estado = "Pendiente"
            if limpiar_fecha:
                c.fecha_descuento_real = None
            pids_tocados.add(c.prestamo_id)
            reabiertas += 1

        # Recalcular estado de cada préstamo afectado:
        for pid in pids_tocados:
            p = Prestamo.query.get(pid)
            n_total = len(p.cuotas or [])
            n_pend = sum(
                1 for cc in p.cuotas if (cc.estado or "Pendiente") == "Pendiente"
            )
            if n_pend == n_total:
                p.estado = "Emitido"
            elif n_pend == 0:
                p.estado = "Cancelado"
            else:
                p.estado = "Amortizado Parcial"

        db.session.commit()
        return jsonify(
            {
                "ok": True,
                "reabiertas": reabiertas,
                "prestamos_afectados": len(pids_tocados),
            }
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Fallo en /api/prestamos/aperturar_mes")
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500
